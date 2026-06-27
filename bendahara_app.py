# =============================================================================
# APLIKASI KEUANGAN TERPADU MI AL MA'ARIF TEMPELSARI
# Role-Based Access via tabel daftar_guru (Supabase)
#
# CARA MENJALANKAN:
#   pip install -r requirements.txt
#   streamlit run bendahara_app.py
# =============================================================================

import base64
import io
import json
import re
import urllib.parse
import uuid
from calendar import monthrange
from datetime import date, datetime

import pandas as pd
import plotly.express as px
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from supabase import Client, create_client

# =============================================================================
# KONFIGURASI
# =============================================================================

SUPABASE_URL = "https://kwrpfcjyavybexkaunsq.supabase.co"
SUPABASE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imt3cnBmY2p5YXZ5YmV4a2F1bnNxIiwicm9sZSI6"
    "ImFub24iLCJpYXQiOjE3ODIyMDAyMzUsImV4cCI6MjA5Nzc3NjIzNX0."
    "T0lojeTGTt9Yv5daouRY2mA74OQsec6E61ZvGJVW7r4"
)

BUCKET_NOTA = "nota_transaksi"
TIPE_FILE_NOTA = ["png", "jpg", "jpeg"]
LOGO_JPG = "logomi.jpg"
LOGO_LEBAR_UTAMA = 180
OFFLINE_LS_KEY = "mi_tempelsari_offline_queue"
REMEMBER_LS_KEY = "mi_tempelsari_remember"

ROLE_GURU_MURNI = "GURU_MURNI"
ROLE_GURU_BENDAHARA = "GURU_BENDAHARA"
ROLE_TU_BENDAHARA = "TU_BENDAHARA"
ROLE_KEPSEK_KOMITE = "KEPSEK_KOMITE"

ROLE_LABEL = {
    ROLE_GURU_MURNI: "Guru Murni",
    ROLE_GURU_BENDAHARA: "Guru Bendahara",
    ROLE_TU_BENDAHARA: "TU Bendahara",
    ROLE_KEPSEK_KOMITE: "Kepsek / Komite",
}

MENU_BY_ROLE = {
    ROLE_GURU_MURNI: ["Absen Kelas", "Iuran Kelas", "Tabungan Kelas", "Data Siswa"],
    ROLE_GURU_BENDAHARA: [
        "Absen Kelas", "Iuran Kelas", "Tabungan Kelas", "Data Siswa",
        "Input Transaksi Bendahara", "Rekap Transaksi",
    ],
    ROLE_TU_BENDAHARA: ["Input Transaksi Bendahara", "Rekap Transaksi"],
    ROLE_KEPSEK_KOMITE: ["Dashboard Eksekutif"],
}

ICON_MENU = {
    "Absen Kelas": "📝", "Iuran Kelas": "💳", "Tabungan Kelas": "🏦",
    "Data Siswa": "👧", "Input Transaksi Bendahara": "💰",
    "Rekap Transaksi": "📊", "Dashboard Eksekutif": "📈",
}

MENU_SLUG = {
    "Absen Kelas": "absen_kelas",
    "Iuran Kelas": "iuran_kelas",
    "Tabungan Kelas": "tabungan_kelas",
    "Data Siswa": "data_siswa",
    "Input Transaksi Bendahara": "input_transaksi",
    "Rekap Transaksi": "rekap_transaksi",
    "Dashboard Eksekutif": "dashboard_eksekutif",
}

MENU_PASTEL = {
    "Absen Kelas": {
        "bg1": "#D8F3DC", "bg2": "#B7E4C7", "teks": "#1B4332",
        "border": "#95D5B2", "shadow": "rgba(27, 67, 50, 0.18)",
    },
    "Iuran Kelas": {
        "bg1": "#FFF9DB", "bg2": "#FFE8A3", "teks": "#5C4A00",
        "border": "#F0C929", "shadow": "rgba(92, 74, 0, 0.16)",
    },
    "Tabungan Kelas": {
        "bg1": "#DBEAFE", "bg2": "#BFDBFE", "teks": "#1E3A5F",
        "border": "#93C5FD", "shadow": "rgba(30, 58, 95, 0.16)",
    },
    "Data Siswa": {
        "bg1": "#FFE5D9", "bg2": "#FFCDB2", "teks": "#7F2D12",
        "border": "#FFB4A2", "shadow": "rgba(127, 45, 18, 0.16)",
    },
    "Input Transaksi Bendahara": {
        "bg1": "#E9D5FF", "bg2": "#D8B4FE", "teks": "#4C1D95",
        "border": "#C084FC", "shadow": "rgba(76, 29, 149, 0.16)",
    },
    "Rekap Transaksi": {
        "bg1": "#CCFBF1", "bg2": "#99F6E4", "teks": "#134E4A",
        "border": "#5EEAD4", "shadow": "rgba(19, 78, 74, 0.16)",
    },
    "Dashboard Eksekutif": {
        "bg1": "#FCE7F3", "bg2": "#FBCFE8", "teks": "#831843",
        "border": "#F9A8D4", "shadow": "rgba(131, 24, 67, 0.16)",
    },
}

DAFTAR_POS_DANA = [
    "Infak Tahunan", "Ngaji Pagi", "Infak Jumat", "Mobil", "Buku LKS",
    "BOS", "Sumbangan Lainnya", "Kantin", "Syukuran", "PIP", "DANANTARA",
]

DAFTAR_KELAS = [
    "IA", "IB", "IIA", "IIB", "IIIA", "IIIB", "IVA", "IVB", "VA", "VB", "VIA", "VIB",
]

POS_TAGIHAN_SISWA = [
    "Infak Tahunan", "Infak Jumat", "Ngaji Pagi", "Mobil", "Buku LKS", "Syukuran", "Seragam",
]

KONFIG_TAGIHAN = {
    "Infak Tahunan": {"tipe": "flat_tahun", "nominal": 200_000},
    "Infak Jumat": {"tipe": "sukarela", "nominal": 0},
    "Ngaji Pagi": {"tipe": "flat_bulan", "nominal": 30_000},
    "Mobil": {"tipe": "kolom", "kolom": "biaya_mobil"},
    "Buku LKS": {"tipe": "kolom", "kolom": "biaya_lks"},
    "Syukuran": {"tipe": "manual"},
    "Seragam": {"tipe": "manual"},
}

KOLOM_TEMPLATE_SISWA = [
    "NO", "Nama Siswa", "NIS", "Nama Ibu", "Alamat",
    "No Whatsapp", "Tanggal Lahir", "Biaya Mobil", "Biaya LKS",
]

STATUS_ABSEN = [".", "S", "I", "A"]

NAMA_BULAN = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April", 5: "Mei", 6: "Juni",
    7: "Juli", 8: "Agustus", 9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}

HARI_LIBUR_TETAP = {
    "01-01", "05-01", "06-01", "03-31", "04-18", "04-19",
    "05-01", "05-29", "06-01", "08-17", "12-25",
}


# =============================================================================
# SUPABASE
# =============================================================================

@st.cache_resource
def koneksi_supabase() -> Client:
    url, key = SUPABASE_URL, SUPABASE_KEY
    try:
        url = st.secrets.get("SUPABASE_URL", url)
        key = st.secrets.get("SUPABASE_KEY", st.secrets.get("SUPABASE_ANON_KEY", key))
    except Exception:
        pass
    try:
        url = st.secrets["supabase"]["url"]
        key = st.secrets["supabase"].get("key") or st.secrets["supabase"].get("anon_key")
    except Exception:
        pass
    return create_client(url, key)


def pastikan_bucket_storage():
    try:
        sb = koneksi_supabase()
        nama = [b.get("name", getattr(b, "name", "")) for b in sb.storage.list_buckets()]
        if BUCKET_NOTA not in nama:
            sb.storage.create_bucket(BUCKET_NOTA, options={"public": True})
    except Exception:
        pass


# =============================================================================
# UTILITAS
# =============================================================================

def format_rupiah(angka) -> str:
    try:
        return f"Rp {float(angka):,.0f}".replace(",", ".")
    except (TypeError, ValueError):
        return "Rp 0"


def indeks_aman(daftar: list, nilai) -> int:
    try:
        return daftar.index(nilai)
    except ValueError:
        return 0


def normalisasi_pos(pos) -> str:
    return str(pos or "").strip().upper()


def pos_ke_label_db(pos_norm: str) -> str:
    for label in DAFTAR_POS_DANA:
        if normalisasi_pos(label) == pos_norm:
            return label
    return pos_norm


def parsing_pos_bendahara(pos_raw):
    if not pos_raw:
        return []
    items = pos_raw if isinstance(pos_raw, list) else re.split(r"[,;|\n]+", str(pos_raw))
    hasil = [pos_ke_label_db(normalisasi_pos(item)) for item in items if normalisasi_pos(item)]
    return sorted(set(hasil))


def bersihkan_whatsapp(nomor) -> str:
    if pd.isna(nomor) or nomor is None:
        return ""
    teks = str(nomor).strip().replace(" ", "").replace("-", "").replace("+", "")
    if teks.startswith("0"):
        teks = "62" + teks[1:]
    elif not teks.startswith("62"):
        teks = "62" + teks
    return teks


def link_whatsapp(nomor: str, pesan: str) -> str:
    nomor_bersih = bersihkan_whatsapp(nomor)
    if not nomor_bersih:
        return ""
    return "https://api.whatsapp.com/send?phone=" + nomor_bersih + "&text=" + urllib.parse.quote(pesan)


def pesan_nota_bayar_jawa(nama_siswa: str, pos_tagihan: str) -> str:
    return (
        f"Matur nuwun sanget, iuran {pos_tagihan} kagem putra/putri panjenengan "
        f"{nama_siswa} sampun dipun tampi dening Guru Kelas. Mugi dados berkah, amin."
    )


def pesan_ultah_boso_jowo(nama_siswa: str) -> str:
    return (
        f"Assalamu'alaikum Wr. Wb.\n\nMatur nuwun sanget dhumateng Bapak/Ibu wali murid.\n"
        f"Kula badhe ngaturaken bilih putra/putri panjenengan *{nama_siswa}* "
        f"sampun nglampahi wilujeng tginggal taun ing dinten niki.\n\n"
        f"Mugiya dipun paringi kesehatan, kebahagiaan, lan kaslametan.\n"
        f"Wilujeng tginggal taun! 🎂\n\nGuru Kelas MI Al Ma'arif Tempelsari"
    )


def tampilkan_tombol_whatsapp(label, nomor, pesan, key=None) -> bool:
    url = link_whatsapp(nomor, pesan)
    if not url:
        st.caption("⚠️ Nomor WhatsApp belum diisi.")
        return False
    st.link_button(label, url=url, use_container_width=True, key=key)
    return True


def muat_ulang_aplikasi():
    st.cache_data.clear()
    st.rerun()


def tampilkan_logo_sidebar():
    try:
        st.image(LOGO_JPG, use_container_width=True)
    except Exception:
        pass


def tampilkan_logo_utama():
    try:
        st.image(LOGO_JPG, width=LOGO_LEBAR_UTAMA)
    except Exception:
        pass


# =============================================================================
# OFFLINE — LocalStorage draft kuning
# =============================================================================

def init_offline_state():
    for k, v in {"offline_queue": [], "offline_draft_baru": False}.items():
        if k not in st.session_state:
            st.session_state[k] = v


def cek_koneksi_internet() -> bool:
    try:
        koneksi_supabase().table("tabel_master_siswa").select("id").limit(1).execute()
        return True
    except Exception:
        return False


def gabung_antrean_dari_url():
    if "offline_data" not in st.query_params:
        return
    try:
        encoded = st.query_params.get("offline_data")
        if isinstance(encoded, list):
            encoded = encoded[0]
        antrean = json.loads(base64.b64decode(encoded).decode("utf-8"))
        if not isinstance(antrean, list):
            return
        id_ada = {x.get("id") for x in st.session_state.offline_queue}
        for item in antrean:
            if item.get("id") not in id_ada:
                st.session_state.offline_queue.append(item)
                id_ada.add(item.get("id"))
        del st.query_params["offline_data"]
    except Exception:
        pass


def _js_escape_json(data) -> str:
    return json.dumps(data).replace("</", "<\\/")


def render_monitor_koneksi_offline():
    html = f"""
    <div id="mi-offline-status" style="font-size:0.85rem;padding:6px 0;text-align:center;">Memeriksa...</div>
    <script>
    (function() {{
        const LS_KEY = "{OFFLINE_LS_KEY}";
        const el = document.getElementById("mi-offline-status");
        function getQueue() {{ try {{ return JSON.parse(localStorage.getItem(LS_KEY)||"[]"); }} catch(e) {{ return []; }} }}
        function updateStatus() {{
            const n = getQueue().length, online = navigator.onLine;
            el.innerHTML = online
                ? (n>0 ? "🟢 Online · <b>"+n+"</b> draft kuning" : "🟢 Online")
                : (n>0 ? "🔴 Offline · <b>"+n+"</b> draft aman" : "🔴 Offline");
            el.style.color = online ? "#1e8449" : "#c0392b";
        }}
        function pushQueue() {{
            const q = getQueue();
            if (!q.length || sessionStorage.getItem("mi_offline_pushed")) return;
            try {{
                const url = new URL(window.top.location.href);
                if (url.searchParams.has("offline_data")) return;
                url.searchParams.set("offline_data", btoa(unescape(encodeURIComponent(JSON.stringify(q)))));
                sessionStorage.setItem("mi_offline_pushed","1");
                window.top.location.href = url.toString();
            }} catch(e) {{}}
        }}
        updateStatus(); pushQueue();
        setInterval(updateStatus, 4000);
        window.addEventListener("online", updateStatus);
        window.addEventListener("offline", updateStatus);
    }})();
    </script>"""
    try:
        components.html(html, height=36)
    except Exception:
        pass


def simpan_ke_localstorage_browser(item: dict):
    html = f"""<script>(function(){{
        try {{
            const KEY="{OFFLINE_LS_KEY}", item={_js_escape_json(item)};
            let q=JSON.parse(localStorage.getItem(KEY)||"[]"); q.push(item);
            localStorage.setItem(KEY,JSON.stringify(q));
            sessionStorage.removeItem("mi_offline_pushed");
        }} catch(e) {{}}
    }})();</script>"""
    try:
        components.html(html, height=0)
    except Exception:
        pass


def bersihkan_localstorage_browser():
    html = f"""<script>try{{localStorage.setItem("{OFFLINE_LS_KEY}","[]");
    sessionStorage.setItem("mi_offline_pushed","1");}}catch(e){{}}</script>"""
    try:
        components.html(html, height=0)
    except Exception:
        pass


def tambah_antrean_offline(tipe: str, payload: dict):
    item = {"id": str(uuid.uuid4()), "type": tipe, "payload": payload,
            "created_at": datetime.now().isoformat()}
    st.session_state.offline_queue.append(item)
    st.session_state.offline_draft_baru = True
    simpan_ke_localstorage_browser(item)


def jumlah_antrean_offline() -> int:
    return len(st.session_state.get("offline_queue", []))


def _simpan_item_antrean(tipe, payload) -> bool:
    if tipe == "pembayaran_siswa":
        return simpan_pembayaran_siswa(**payload)
    if tipe == "tabungan_siswa":
        return simpan_tabungan(**payload)
    if tipe == "transaksi_bendahara":
        return simpan_transaksi_bendahara(**payload)
    if tipe == "absen_siswa":
        return upsert_absen_siswa(**payload)
    return False


def sinkronkan_antrean_offline() -> tuple[int, int]:
    sukses = gagal = 0
    sisa = []
    for item in list(st.session_state.get("offline_queue", [])):
        if _simpan_item_antrean(item.get("type"), item.get("payload", {})):
            sukses += 1
        else:
            gagal += 1
            sisa.append(item)
    st.session_state.offline_queue = sisa
    if sukses and not sisa:
        st.session_state.offline_draft_baru = False
    return sukses, gagal


def simpan_dengan_offline(tipe, payload, pesan_sukses, meta_wa=None):
    if cek_koneksi_internet() and _simpan_item_antrean(tipe, payload):
        if meta_wa and meta_wa.get("no_wa"):
            st.session_state.wa_draft_link = link_whatsapp(
                meta_wa["no_wa"], pesan_nota_bayar_jawa(meta_wa["nama_siswa"], meta_wa["pos_tagihan"]))
            st.session_state.wa_draft_label = f"📱 Kirim Nota — {meta_wa['nama_siswa']}"
        else:
            st.session_state.wa_draft_link = st.session_state.wa_draft_label = None
        st.session_state.flash_pesan = pesan_sukses
        muat_ulang_aplikasi()
        return
    tambah_antrean_offline(tipe, payload)
    st.warning("⚠️ **Draft kuning** — internet terputus. Sinkronkan lewat sidebar saat online.")


# =============================================================================
# AUTH — daftar_guru RBAC
# Kolom Supabase: username, password (plain), role, kelas_diampu
# =============================================================================

KOLOM_DAFTAR_GURU = "id,username,password,role,kelas_diampu"


def normalisasi_peran(role):
    """Normalisasi nilai kolom `role` ke konstanta internal."""
    if not role:
        return None
    p = str(role).strip().upper().replace(" ", "_").replace("-", "_")
    return {
        "GURU": ROLE_GURU_MURNI, "GURU_KELAS": ROLE_GURU_MURNI, "GURU_MURNI": ROLE_GURU_MURNI,
        "GURU_BENDAHARA": ROLE_GURU_BENDAHARA, "BENDAHARA_GURU": ROLE_GURU_BENDAHARA,
        "TU": ROLE_TU_BENDAHARA, "TU_BENDAHARA": ROLE_TU_BENDAHARA,
        "KEPSEK": ROLE_KEPSEK_KOMITE, "KOMITE": ROLE_KEPSEK_KOMITE,
        "KEPSEK_KOMITE": ROLE_KEPSEK_KOMITE, "KEPALA_SEKOLAH": ROLE_KEPSEK_KOMITE,
    }.get(p, p)


def init_auth_state():
    defaults = {
        "authenticated": False, "remember_me": False, "user_id": None,
        "nama_guru": None, "username": None, "peran_aktif": None,
        "kelas_guru": None, "pos_bendahara": [], "halaman_aktif": None,
        "flash_pesan": None, "wa_draft_link": None, "wa_draft_label": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def ambil_guru_dari_supabase(username):
    """Ambil satu baris dari daftar_guru — cocokkan hanya kolom `username`."""
    username = str(username or "").strip()
    if not username:
        return None
    try:
        res = (
            koneksi_supabase()
            .table("daftar_guru")
            .select(KOLOM_DAFTAR_GURU)
            .eq("username", username)
            .limit(1)
            .execute()
        )
        return res.data[0] if res.data else None
    except Exception as e:
        st.error(f"Gagal membaca daftar_guru: {e}")
        return None


def ambil_guru_by_id(user_id):
    try:
        res = (
            koneksi_supabase()
            .table("daftar_guru")
            .select(KOLOM_DAFTAR_GURU)
            .eq("id", user_id)
            .limit(1)
            .execute()
        )
        return res.data[0] if res.data else None
    except Exception:
        return None


def verifikasi_password(password_input, guru) -> bool:
    """Cocokkan password input dengan kolom `password` (teks biasa, tanpa hash)."""
    return str(password_input or "") == str(guru.get("password") or "")


def _set_session_guru(guru, ingat_saya=False):
    peran = normalisasi_peran(guru.get("role"))
    st.session_state.authenticated = True
    st.session_state.remember_me = ingat_saya
    st.session_state.user_id = guru.get("id")
    st.session_state.username = guru.get("username")
    st.session_state.nama_guru = guru.get("username")
    st.session_state.peran_aktif = peran
    st.session_state.kelas_guru = guru.get("kelas_diampu")
    st.session_state.pos_bendahara = parsing_pos_bendahara(guru.get("pos_bendahara"))
    menu = menu_berdasarkan_peran(peran)
    st.session_state.halaman_aktif = menu[0] if menu else None


def proses_login_daftar_guru(username, password, ingat_saya=False):
    username = str(username or "").strip()
    guru = ambil_guru_dari_supabase(username)
    if not guru:
        return False, f"Username '{username}' tidak ditemukan di daftar_guru."
    if not verifikasi_password(password, guru):
        return False, "Password salah."
    role_norm = normalisasi_peran(guru.get("role"))
    if role_norm not in MENU_BY_ROLE:
        return False, f"Role '{guru.get('role')}' belum dikenali sistem."
    _set_session_guru(guru, ingat_saya)
    if ingat_saya:
        data = base64.b64encode(json.dumps({"user_id": guru["id"]}).encode()).decode()
        components.html(f'<script>try{{localStorage.setItem("{REMEMBER_LS_KEY}","{data}");}}catch(e){{}}</script>', height=0)
    else:
        components.html(f'<script>try{{localStorage.removeItem("{REMEMBER_LS_KEY}");}}catch(e){{}}</script>', height=0)
    return True, f"Selamat datang, {guru.get('username')}!"


def proses_logout():
    components.html(f'<script>try{{localStorage.removeItem("{REMEMBER_LS_KEY}");}}catch(e){{}}</script>', height=0)
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    init_auth_state()
    init_offline_state()
    muat_ulang_aplikasi()


def ubah_password_mandiri(password_lama, password_baru):
    if not st.session_state.get("user_id"):
        return False, "Silakan login dulu."
    if len(password_baru) < 6:
        return False, "Password baru minimal 6 karakter."
    guru = ambil_guru_dari_supabase(st.session_state.get("username"))
    if not guru or not verifikasi_password(password_lama, guru):
        return False, "Password lama salah."
    koneksi_supabase().table("daftar_guru").update(
        {"password": password_baru}
    ).eq("id", st.session_state.user_id).execute()
    return True, "Password berhasil diubah."


def menu_berdasarkan_peran(peran):
    return MENU_BY_ROLE.get(normalisasi_peran(peran), [])


def pastikan_akses_halaman(halaman):
    if halaman not in menu_berdasarkan_peran(st.session_state.get("peran_aktif")):
        st.error("⛔ Anda tidak memiliki akses ke halaman ini.")
        st.stop()


def pos_input_tersedia():
    peran = st.session_state.get("peran_aktif")
    if peran == ROLE_GURU_BENDAHARA:
        pos = st.session_state.get("pos_bendahara", [])
        return pos if pos else DAFTAR_POS_DANA
    if peran == ROLE_TU_BENDAHARA:
        return ["DANANTARA"]
    return []


def boleh_input_transaksi_bendahara(pos):
    peran = st.session_state.get("peran_aktif")
    if peran == ROLE_GURU_BENDAHARA:
        daftar = st.session_state.get("pos_bendahara", [])
        return pos in daftar if daftar else pos in DAFTAR_POS_DANA
    if peran == ROLE_TU_BENDAHARA:
        return normalisasi_pos(pos) == "DANANTARA"
    return False


def punya_akses_kelas():
    """Guru murni & guru bendahara mengampu kelas dari kolom kelas_diampu."""
    return st.session_state.get("peran_aktif") in (ROLE_GURU_MURNI, ROLE_GURU_BENDAHARA)


def boleh_lihat_rekap_semua_pos():
    return st.session_state.get("peran_aktif") in (
        ROLE_GURU_BENDAHARA, ROLE_TU_BENDAHARA, ROLE_KEPSEK_KOMITE)


def coba_auto_login_dari_remember():
    if st.session_state.get("authenticated"):
        return
    if "auto_user" in st.query_params:
        try:
            guru = ambil_guru_by_id(int(st.query_params.get("auto_user")))
            if guru:
                _set_session_guru(guru, ingat_saya=True)
            del st.query_params["auto_user"]
        except Exception:
            pass
        return
    html = f"""<script>(function(){{
        if(sessionStorage.getItem("mi_auto_login_done"))return;
        try{{
            const raw=localStorage.getItem("{REMEMBER_LS_KEY}");
            if(!raw)return;
            const d=JSON.parse(atob(raw));
            if(!d.user_id)return;
            const url=new URL(window.top.location.href);
            if(url.searchParams.has("auto_user"))return;
            url.searchParams.set("auto_user",d.user_id);
            sessionStorage.setItem("mi_auto_login_done","1");
            window.top.location.href=url.toString();
        }}catch(e){{}}
    }})();</script>"""
    try:
        components.html(html, height=0)
    except Exception:
        pass


def render_form_ubah_password():
    with st.expander("🔑 Ubah Password Mandiri"):
        with st.form("form_ubah_password"):
            lama = st.text_input("Password Lama", type="password")
            baru = st.text_input("Password Baru", type="password")
            konfirm = st.text_input("Konfirmasi Password Baru", type="password")
            if st.form_submit_button("Simpan Password Baru", use_container_width=True):
                if baru != konfirm:
                    st.error("Konfirmasi password tidak cocok.")
                else:
                    ok, pesan = ubah_password_mandiri(lama, baru)
                    st.success(pesan) if ok else st.error(pesan)


# =============================================================================
# KEUANGAN BENDAHARA — tabel_transaksi_bendahara (legacy)
# =============================================================================

def ambil_transaksi_bendahara() -> pd.DataFrame:
    try:
        res = koneksi_supabase().table("tabel_transaksi_bendahara").select("*").order("tanggal", desc=True).execute()
        if not res.data:
            return pd.DataFrame()
        df = pd.DataFrame(res.data)
        df["nominal"] = pd.to_numeric(df["nominal"], errors="coerce").fillna(0)
        df["tanggal"] = pd.to_datetime(df["tanggal"], errors="coerce")
        if "foto_nota" not in df.columns:
            df["foto_nota"] = None
        return df
    except Exception as e:
        st.error(f"Gagal ambil transaksi: {e}")
        return pd.DataFrame()


def simpan_transaksi_bendahara(**data) -> bool:
    try:
        koneksi_supabase().table("tabel_transaksi_bendahara").insert(data).execute()
        return True
    except Exception as e:
        st.error(f"Gagal simpan transaksi: {e}")
        return False


def unggah_foto_nota(file_bytes, nama_file, tipe_file):
    try:
        sb = koneksi_supabase()
        pastikan_bucket_storage()
        ext = nama_file.rsplit(".", 1)[-1].lower()
        path = f"nota/{datetime.now():%Y%m%d_%H%M%S}_{uuid.uuid4().hex[:8]}.{ext}"
        sb.storage.from_(BUCKET_NOTA).upload(
            path=path, file=file_bytes,
            file_options={"content-type": tipe_file or "image/jpeg", "upsert": "false"})
        return sb.storage.from_(BUCKET_NOTA).get_public_url(path)
    except Exception as e:
        st.error(f"Gagal unggah foto: {e}")
        return None


def ringkasan_bendahara(df):
    if df.empty:
        return {"pemasukan": 0.0, "pengeluaran": 0.0, "saldo": 0.0}
    masuk = df.loc[df["jenis_transaksi"] == "Pemasukan", "nominal"].sum()
    keluar = df.loc[df["jenis_transaksi"] == "Pengeluaran", "nominal"].sum()
    return {"pemasukan": float(masuk), "pengeluaran": float(keluar), "saldo": float(masuk - keluar)}


def filter_bulan_tahun(df, bulan, tahun):
    if df.empty:
        return df.copy()
    m = (df["tanggal"].dt.month == bulan) & (df["tanggal"].dt.year == tahun)
    return df.loc[m].copy()


def laporan_per_pos_dana(df_bulan):
    hasil = {}
    if df_bulan.empty:
        return hasil
    total_masuk = df_bulan.loc[df_bulan["jenis_transaksi"] == "Pemasukan", "nominal"].sum()
    for pos in DAFTAR_POS_DANA:
        df_pos = df_bulan[df_bulan["pos_dana"] == pos]
        if df_pos.empty:
            continue
        masuk = df_pos.loc[df_pos["jenis_transaksi"] == "Pemasukan", "nominal"].sum()
        keluar = df_pos.loc[df_pos["jenis_transaksi"] == "Pengeluaran", "nominal"].sum()
        hasil[pos] = {
            "detail": df_pos.sort_values("tanggal"),
            "pemasukan": float(masuk), "pengeluaran": float(keluar),
            "saldo": float(masuk - keluar),
            "persen": round((masuk / total_masuk * 100) if total_masuk else 0, 1),
        }
    return hasil


def buat_excel_laporan_pos(bulan, tahun, ringkasan, laporan_pos):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        pd.DataFrame([
            ["Laporan MI Al Ma'arif Tempelsari"], ["Periode", f"{NAMA_BULAN[bulan]} {tahun}"],
            ["Pemasukan", ringkasan["pemasukan"]], ["Pengeluaran", ringkasan["pengeluaran"]],
            ["Saldo", ringkasan["saldo"]],
        ]).to_excel(w, sheet_name="Ringkasan", index=False, header=False)
        for pos, data in laporan_pos.items():
            df_out = data["detail"].copy()
            df_out["tanggal"] = df_out["tanggal"].dt.strftime("%d/%m/%Y")
            df_out[["tanggal", "nama_bendahara", "jenis_transaksi", "nominal", "keterangan"]].to_excel(
                w, sheet_name=pos[:31], index=False)
    return buf.getvalue()


def dataframe_to_excel_bytes(df, sheet_name="Rekap"):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=sheet_name)
    return buf.getvalue()


def hitung_saldo_konsolidasi_sekolah():
    return ringkasan_bendahara(ambil_transaksi_bendahara())["saldo"]


def hitung_total_pemasukan():
    return ringkasan_bendahara(ambil_transaksi_bendahara())["pemasukan"]


def hitung_total_pengeluaran():
    return ringkasan_bendahara(ambil_transaksi_bendahara())["pengeluaran"]


def saldo_per_pos_dana():
    df = ambil_transaksi_bendahara()
    if df.empty:
        return pd.DataFrame(columns=["Pos Dana", "Pemasukan", "Pengeluaran", "Saldo"])
    baris = []
    for pos in DAFTAR_POS_DANA:
        d = df[df["pos_dana"] == pos]
        if d.empty:
            continue
        masuk = d.loc[d["jenis_transaksi"] == "Pemasukan", "nominal"].sum()
        keluar = d.loc[d["jenis_transaksi"] == "Pengeluaran", "nominal"].sum()
        baris.append({"Pos Dana": pos, "Pemasukan": float(masuk),
                      "Pengeluaran": float(keluar), "Saldo": float(masuk - keluar)})
    return pd.DataFrame(baris)


def jenis_transaksi_dari_kode(kode):
    return "Pemasukan" if str(kode).upper() in ("IN", "MASUK", "PEMASUKAN") else "Pengeluaran"


# =============================================================================
# SISWA — iuran, tabungan, absen
# =============================================================================

def ambil_siswa(kelas=None):
    try:
        q = koneksi_supabase().table("tabel_master_siswa").select("*").order("no")
        if kelas:
            q = q.eq("kelas", kelas)
        res = q.execute()
        if not res.data:
            return pd.DataFrame()
        df = pd.DataFrame(res.data)
        for kol in ["biaya_mobil", "biaya_lks"]:
            df[kol] = pd.to_numeric(df.get(kol, 0), errors="coerce").fillna(0)
        if "tanggal_lahir" in df.columns:
            df["tanggal_lahir"] = pd.to_datetime(df["tanggal_lahir"], errors="coerce")
        return df
    except Exception as e:
        st.error(f"Gagal ambil siswa: {e}")
        return pd.DataFrame()


def hapus_siswa_per_kelas(kelas):
    try:
        koneksi_supabase().table("tabel_master_siswa").delete().eq("kelas", kelas).execute()
        return True
    except Exception:
        return False


def insert_siswa_bulk(records):
    try:
        if records:
            koneksi_supabase().table("tabel_master_siswa").insert(records).execute()
        return True
    except Exception:
        return False


def buat_template_excel_siswa():
    df = pd.DataFrame(columns=KOLOM_TEMPLATE_SISWA)
    df.loc[0] = [1, "Contoh Nama Siswa", "12345", "Contoh Nama Ibu", "Alamat",
                 "6281234567890", "2015-06-15", 50000, 75000]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Template Siswa", index=False)
    return buf.getvalue()


def proses_upload_excel_siswa(file_bytes, kelas):
    try:
        df = pd.read_excel(io.BytesIO(file_bytes))
        df.columns = [str(c).strip() for c in df.columns]
        if "Nama Siswa" not in df.columns:
            return False, "Kolom 'Nama Siswa' wajib ada."
        records = []
        for _, row in df.iterrows():
            nama = str(row.get("Nama Siswa", "")).strip()
            if not nama or nama.lower() == "contoh nama siswa":
                continue
            rec = {
                "no": int(row.get("NO", 0) or 0), "nama_siswa": nama,
                "nis": str(row.get("NIS", "") or ""), "nama_ibu": str(row.get("Nama Ibu", "") or ""),
                "alamat": str(row.get("Alamat", "") or ""),
                "no_whatsapp": bersihkan_whatsapp(row.get("No Whatsapp", "")),
                "biaya_mobil": float(row.get("Biaya Mobil", 0) or 0),
                "biaya_lks": float(row.get("Biaya LKS", 0) or 0), "kelas": kelas,
            }
            tgl = row.get("Tanggal Lahir")
            if pd.notna(tgl) and str(tgl).strip():
                try:
                    rec["tanggal_lahir"] = pd.to_datetime(tgl).strftime("%Y-%m-%d")
                except Exception:
                    pass
            records.append(rec)
        if not records:
            return False, "Tidak ada data valid."
        if not hapus_siswa_per_kelas(kelas) or not insert_siswa_bulk(records):
            return False, "Gagal menyimpan."
        return True, f"Berhasil impor {len(records)} siswa kelas {kelas}."
    except Exception as e:
        return False, f"Error: {e}"


def siswa_ultah_hari_ini(kelas):
    df = ambil_siswa(kelas)
    if df.empty or "tanggal_lahir" not in df.columns:
        return pd.DataFrame()
    h = date.today()
    mask = df["tanggal_lahir"].notna() & (df["tanggal_lahir"].dt.month == h.month) & (df["tanggal_lahir"].dt.day == h.day)
    return df.loc[mask]


def ambil_pembayaran_siswa(kelas=None):
    try:
        q = koneksi_supabase().table("tabel_pembayaran_siswa").select("*").order("tanggal", desc=True)
        if kelas:
            q = q.eq("kelas", kelas)
        res = q.execute()
        if not res.data:
            return pd.DataFrame()
        df = pd.DataFrame(res.data)
        df["nominal"] = pd.to_numeric(df["nominal"], errors="coerce").fillna(0)
        df["tanggal"] = pd.to_datetime(df["tanggal"], errors="coerce")
        return df
    except Exception:
        return pd.DataFrame()


def simpan_pembayaran_siswa(**data):
    try:
        koneksi_supabase().table("tabel_pembayaran_siswa").insert(data).execute()
        return True
    except Exception:
        return False


def ambil_target_manual(kelas, tahun):
    try:
        res = koneksi_supabase().table("tabel_target_manual").select("*").eq(
            "kelas", kelas).eq("periode_tahun", tahun).execute()
        return pd.DataFrame(res.data or [])
    except Exception:
        return pd.DataFrame()


def hitung_target_siswa(siswa, pos, target_manual, bulan, tahun):
    cfg = KONFIG_TAGIHAN.get(pos, {})
    tipe = cfg.get("tipe")
    if tipe in ("flat_tahun", "flat_bulan"):
        return float(cfg["nominal"])
    if tipe == "sukarela":
        return 0.0
    if tipe == "kolom":
        return float(siswa.get(cfg["kolom"], 0) or 0)
    if tipe == "manual" and not target_manual.empty:
        baris = target_manual[target_manual["pos_tagihan"] == pos]
        return float(baris.iloc[0]["nominal"]) if not baris.empty else 0.0
    return 0.0


def total_bayar_siswa_pos(df_bayar, siswa_id, pos, bulan, tahun):
    if df_bayar.empty:
        return 0.0
    mask = ((df_bayar["siswa_id"] == siswa_id) & (df_bayar["pos_tagihan"] == pos)
            & (df_bayar["periode_bulan"] == bulan) & (df_bayar["periode_tahun"] == tahun))
    return float(df_bayar.loc[mask, "nominal"].sum())


def status_pelunasan(persen, target, bayar):
    if target <= 0:
        return "Sukarela" if bayar > 0 else "Bebas"
    if persen >= 100:
        return "✅ Lunas"
    if bayar <= 0:
        return "❌ Belum Bayar"
    return f"🟡 Cicil ({persen:.0f}%)"


def buat_rekap_tunggakan(kelas, bulan, tahun):
    df_siswa = ambil_siswa(kelas)
    if df_siswa.empty:
        return pd.DataFrame()
    df_bayar = ambil_pembayaran_siswa(kelas)
    target_manual = ambil_target_manual(kelas, tahun)
    baris = []
    for _, siswa in df_siswa.iterrows():
        for pos in POS_TAGIHAN_SISWA:
            target = hitung_target_siswa(siswa, pos, target_manual, bulan, tahun)
            bayar = total_bayar_siswa_pos(df_bayar, siswa["id"], pos, bulan, tahun)
            if target <= 0 and pos != "Infak Jumat":
                continue
            persen = min(100.0, (bayar / target * 100)) if target else (100.0 if bayar > 0 else 0.0)
            baris.append({
                "Nama Siswa": siswa["nama_siswa"], "Pos Tagihan": pos,
                "Target": target, "Terbayar": bayar, "Sisa": max(0.0, target - bayar),
                "Persentase": round(persen, 1),
                "Status": status_pelunasan(persen, target, bayar),
                "No Whatsapp": siswa.get("no_whatsapp", ""),
            })
    return pd.DataFrame(baris)


def metrik_pelunasan_sekolah(bulan, tahun):
    total_target = total_bayar = 0.0
    per_kelas = []
    for kelas in DAFTAR_KELAS:
        rekap = buat_rekap_tunggakan(kelas, bulan, tahun)
        if rekap.empty:
            continue
        t_target, t_bayar = rekap["Target"].sum(), rekap["Terbayar"].sum()
        total_target += t_target
        total_bayar += t_bayar
        per_kelas.append({"Kelas": kelas, "Target": t_target, "Terbayar": t_bayar,
                          "Persentase": round((t_bayar / t_target * 100) if t_target else 0, 1)})
    return {
        "persen_sekolah": round((total_bayar / total_target * 100) if total_target else 0, 1),
        "per_kelas": pd.DataFrame(per_kelas),
    }


def ambil_tabungan(kelas=None):
    try:
        q = koneksi_supabase().table("tabel_tabungan_siswa").select("*").order("tanggal", desc=True)
        if kelas:
            q = q.eq("kelas", kelas)
        res = q.execute()
        if not res.data:
            return pd.DataFrame()
        df = pd.DataFrame(res.data)
        df["nominal"] = pd.to_numeric(df["nominal"], errors="coerce").fillna(0)
        return df
    except Exception:
        return pd.DataFrame()


def simpan_tabungan(**data):
    try:
        koneksi_supabase().table("tabel_tabungan_siswa").insert(data).execute()
        return True
    except Exception:
        return False


def rekap_tabungan_kelas(kelas):
    df_s, df_t = ambil_siswa(kelas), ambil_tabungan(kelas)
    if df_s.empty:
        return pd.DataFrame()
    return pd.DataFrame([{
        "Nama": s["nama_siswa"],
        "Nominal": float(df_t.loc[df_t["siswa_id"] == s["id"], "nominal"].sum()) if not df_t.empty else 0.0,
    } for _, s in df_s.iterrows()])


def ambil_absen_kelas(kelas, bulan, tahun):
    try:
        akhir = monthrange(tahun, bulan)[1]
        res = koneksi_supabase().table("tabel_absen_siswa").select("*").eq("kelas", kelas).gte(
            "tanggal", f"{tahun}-{bulan:02d}-01").lte("tanggal", f"{tahun}-{bulan:02d}-{akhir}").execute()
        return pd.DataFrame(res.data or [])
    except Exception:
        return pd.DataFrame()


def upsert_absen_siswa(**data):
    try:
        sb = koneksi_supabase()
        ex = sb.table("tabel_absen_siswa").select("id").eq(
            "siswa_id", data["siswa_id"]).eq("tanggal", data["tanggal"]).limit(1).execute()
        if ex.data:
            sb.table("tabel_absen_siswa").update(data).eq("id", ex.data[0]["id"]).execute()
        else:
            sb.table("tabel_absen_siswa").insert(data).execute()
        return True
    except Exception as e:
        st.error(f"Gagal simpan absen: {e}")
        return False


def _adalah_libur(tgl):
    return tgl.weekday() >= 5 or tgl.strftime("%m-%d") in HARI_LIBUR_TETAP


def buat_excel_rekap_absen(kelas, bulan, tahun, semester=False):
    df_siswa = ambil_siswa(kelas)
    if df_siswa.empty:
        return b""
    if semester:
        bln_awal, bln_akhir = (1, 6) if bulan <= 6 else (7, 12)
        label_periode = f"Semester {'1' if bln_awal == 1 else '2'} {tahun}"
    else:
        bln_awal = bln_akhir = bulan
        label_periode = f"{NAMA_BULAN[bulan]} {tahun}"
    tanggal_list = [date(tahun, b, h) for b in range(bln_awal, bln_akhir + 1)
                    for h in range(1, monthrange(tahun, b)[1] + 1)]
    lookup = {}
    for b in range(bln_awal, bln_akhir + 1):
        part = ambil_absen_kelas(kelas, b, tahun)
        for _, r in part.iterrows() if not part.empty else []:
            lookup[(r["siswa_id"], str(r["tanggal"])[:10])] = r.get("status", ".")
    wb = Workbook()
    ws = wb.active
    ws.title = "Absen"
    ws.cell(1, 1, f"Rekap Absen Kelas {kelas} — {label_periode}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(tanggal_list) + 1)
    ws["A1"].font = Font(bold=True, size=12)
    ws.cell(2, 1, "Nama Siswa")
    merah = PatternFill(start_color="F5C6C6", end_color="F5C6C6", fill_type="solid")
    libur_cols = []
    for i, tgl in enumerate(tanggal_list, start=2):
        c = ws.cell(2, i, tgl.strftime("%d/%m") if semester else str(tgl.day))
        c.alignment = Alignment(textRotation=90, horizontal="center", vertical="bottom")
        if _adalah_libur(tgl):
            libur_cols.append(i)
            c.fill = merah
    for ri, (_, siswa) in enumerate(df_siswa.iterrows(), start=3):
        ws.cell(ri, 1, siswa["nama_siswa"])
        for i, tgl in enumerate(tanggal_list, start=2):
            cell = ws.cell(ri, i, lookup.get((siswa["id"], tgl.isoformat()), "."))
            cell.alignment = Alignment(horizontal="center")
            if i in libur_cols:
                cell.fill = merah
    ws.column_dimensions["A"].width = 28
    for i in range(2, len(tanggal_list) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 4
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# UI
# =============================================================================

def inject_css_global():
    st.markdown("""<style>
    .sub-judul{text-align:center;color:#566573;margin-bottom:1rem;}
    .draft-kuning{background:#FFF9DB;border:1px solid #F0C929;border-radius:10px;padding:10px 14px;margin-bottom:12px;}
    span.menu-anchor{display:none;}
    </style>""", unsafe_allow_html=True)


def inject_css_menu_utama():
    """Warna pastel per menu + bayangan tipis + sudut melengkung pada tombol grid."""
    rules = []
    for nama, g in MENU_PASTEL.items():
        slug = MENU_SLUG.get(nama, nama.lower().replace(" ", "_"))
        sel = f'span#menu-anchor-{slug} ~ div[data-testid="stButton"] > button'
        rules.append(f"""
        {sel}{{
            background:linear-gradient(135deg,{g["bg1"]},{g["bg2"]})!important;
            color:{g["teks"]}!important;
            border:2px solid {g["border"]}!important;
            border-radius:18px!important;
            box-shadow:0 4px 14px {g["shadow"]}!important;
            min-height:72px!important;
            font-size:17px!important;
            font-weight:700!important;
            transition:transform .15s ease,box-shadow .15s ease,filter .15s ease;
        }}
        {sel}:hover{{
            filter:brightness(.97)!important;
            box-shadow:0 6px 18px {g["shadow"]}!important;
            transform:translateY(-2px)!important;
            color:{g["teks"]}!important;
            border-color:{g["border"]}!important;
        }}
        {sel}:focus,{sel}:active{{
            background:linear-gradient(135deg,{g["bg1"]},{g["bg2"]})!important;
            color:{g["teks"]}!important;
            border-color:{g["border"]}!important;
            box-shadow:0 3px 10px {g["shadow"]}!important;
        }}""")
    st.markdown(f"<style>{''.join(rules)}</style>", unsafe_allow_html=True)


def render_notif_ultah(kelas):
    ultah = siswa_ultah_hari_ini(kelas)
    if ultah.empty:
        return
    st.markdown("##### 🎂 Ultah Hari Ini")
    for _, s in ultah.iterrows():
        st.success(f"🎉 **{s['nama_siswa']}** — Wilujeng tginggal taun!")
        tampilkan_tombol_whatsapp(f"📱 WA Ultah — {s['nama_siswa']}", s.get("no_whatsapp", ""),
                                    pesan_ultah_boso_jowo(s["nama_siswa"]), key=f"wa_ultah_{s['id']}")


def render_grid_menu_utama():
    peran = st.session_state.get("peran_aktif")
    if peran == ROLE_KEPSEK_KOMITE:
        return
    menu_aktif = menu_berdasarkan_peran(peran)
    inject_css_menu_utama()
    st.subheader("Menu Utama")
    n_col = 4 if peran in (ROLE_GURU_MURNI, ROLE_GURU_BENDAHARA) else 3
    for i in range(0, len(menu_aktif), n_col):
        cols = st.columns(n_col, gap="large")
        for idx, nama in enumerate(menu_aktif[i:i + n_col]):
            slug = MENU_SLUG.get(nama, nama.lower().replace(" ", "_"))
            with cols[idx]:
                st.markdown(
                    f'<span id="menu-anchor-{slug}" class="menu-anchor"></span>',
                    unsafe_allow_html=True,
                )
                if st.button(
                    f"{ICON_MENU.get(nama, '📌')}  {nama}",
                    key=f"ubin_{slug}",
                    use_container_width=True,
                ):
                    st.session_state.halaman_aktif = nama
                    muat_ulang_aplikasi()


def render_banner_draft_kuning():
    if st.session_state.get("offline_draft_baru") or st.session_state.get("offline_queue"):
        n = len(st.session_state.get("offline_queue", []))
        st.markdown(f'<div class="draft-kuning">📦 <b>{n}</b> draft kuning — sinkronkan via sidebar saat online.</div>',
                    unsafe_allow_html=True)


def _kelas_aktif():
    return st.session_state.get("kelas_guru") or DAFTAR_KELAS[0]


# =============================================================================
# HALAMAN
# =============================================================================

def halaman_absen_kelas():
    pastikan_akses_halaman("Absen Kelas")
    kelas = _kelas_aktif()
    st.header(f"📝 Absen Kelas {kelas}")
    st.caption("Status: **.** Hadir | **S** Sakit | **I** Izin | **A** Alpha")
    render_notif_ultah(kelas)
    c1, c2, c3 = st.columns(3)
    with c1:
        tgl = st.date_input("Tanggal", value=date.today(), key="absen_tgl")
    with c2:
        bln = st.selectbox("Bulan rekap", list(NAMA_BULAN.keys()), format_func=lambda x: NAMA_BULAN[x],
                           index=date.today().month - 1)
    with c3:
        thn = st.selectbox("Tahun", list(range(2024, date.today().year + 2)),
                           index=list(range(2024, date.today().year + 2)).index(date.today().year))
    df_siswa = ambil_siswa(kelas)
    if df_siswa.empty:
        st.warning("Belum ada siswa.")
        return
    st.subheader(f"Input — {tgl.strftime('%d/%m/%Y')}")
    for _, siswa in df_siswa.iterrows():
        c_n, c_s, c_b = st.columns([3, 1, 1])
        with c_n:
            st.write(f"**{siswa['nama_siswa']}**")
        with c_s:
            status = st.selectbox("s", STATUS_ABSEN, key=f"abs_{siswa['id']}_{tgl}", label_visibility="collapsed")
        with c_b:
            if st.button("💾", key=f"sv_{siswa['id']}_{tgl}"):
                payload = {"siswa_id": int(siswa["id"]), "kelas": kelas, "tanggal": tgl.isoformat(),
                           "status": status, "guru_id": st.session_state.get("user_id")}
                if cek_koneksi_internet() and upsert_absen_siswa(**payload):
                    st.session_state.flash_pesan = f"Absen {siswa['nama_siswa']} tersimpan."
                    muat_ulang_aplikasi()
                else:
                    simpan_dengan_offline("absen_siswa", payload, "Absen (draft)")
    st.markdown("---")
    x1, x2 = buat_excel_rekap_absen(kelas, bln, thn, False), buat_excel_rekap_absen(kelas, bln, thn, True)
    if x1:
        st.download_button("📊 Rekap Bulanan", x1, f"Absen_{kelas}_{NAMA_BULAN[bln]}_{thn}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    if x2:
        st.download_button("📊 Rekap Semester", x2, f"Absen_{kelas}_Sem_{thn}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)


def halaman_iuran_kelas():
    pastikan_akses_halaman("Iuran Kelas")
    kelas = _kelas_aktif()
    st.header(f"💳 Iuran Kelas {kelas}")
    render_notif_ultah(kelas)
    if st.session_state.get("wa_draft_link"):
        st.link_button(st.session_state.get("wa_draft_label", "📱 WA Nota"),
                       url=st.session_state.wa_draft_link, use_container_width=True)
        if st.button("✖ Tutup WA"):
            st.session_state.wa_draft_link = None
            st.rerun()
    bulan = st.selectbox("Bulan", list(NAMA_BULAN.keys()), format_func=lambda x: NAMA_BULAN[x],
                         index=date.today().month - 1)
    tahun = st.selectbox("Tahun", list(range(2024, date.today().year + 2)),
                         index=list(range(2024, date.today().year + 2)).index(date.today().year))
    df_siswa = ambil_siswa(kelas)
    if df_siswa.empty:
        st.warning("Belum ada siswa.")
        return
    map_s = {f"{r['nama_siswa']} (NIS:{r.get('nis','-')})": r for _, r in df_siswa.iterrows()}
    siswa = map_s[st.selectbox("Siswa", list(map_s.keys()))]
    pos = st.selectbox("Pos Tagihan", POS_TAGIHAN_SISWA)
    target = hitung_target_siswa(siswa, pos, ambil_target_manual(kelas, tahun), bulan, tahun)
    sudah = total_bayar_siswa_pos(ambil_pembayaran_siswa(kelas), siswa["id"], pos, bulan, tahun)
    st.info(f"Target **{format_rupiah(target)}** | Sudah **{format_rupiah(sudah)}** | Sisa **{format_rupiah(max(0,target-sudah))}**")
    with st.form("bayar"):
        nominal = st.number_input("Nominal", min_value=0, step=1000)
        tgl = st.date_input("Tanggal", value=date.today())
        ket = st.text_input("Keterangan")
        if st.form_submit_button("💾 Simpan", type="primary") and nominal > 0:
            payload = {"siswa_id": int(siswa["id"]), "pos_tagihan": pos, "nominal": float(nominal),
                       "tanggal": tgl.isoformat(), "kelas": kelas, "keterangan": ket,
                       "periode_bulan": bulan, "periode_tahun": tahun}
            meta = {"nama_siswa": siswa["nama_siswa"], "pos_tagihan": pos, "no_wa": siswa.get("no_whatsapp", "")}
            simpan_dengan_offline("pembayaran_siswa", payload, "Pembayaran tersimpan!", meta)
    rekap = buat_rekap_tunggakan(kelas, bulan, tahun)
    if not rekap.empty:
        st.subheader("Rekap Iuran")
        t = rekap.copy()
        for kol in ("Target", "Terbayar", "Sisa"):
            t[kol] = t[kol].apply(format_rupiah)
        st.dataframe(t, use_container_width=True, hide_index=True)


def halaman_tabungan_kelas():
    pastikan_akses_halaman("Tabungan Kelas")
    kelas = _kelas_aktif()
    st.header(f"🏦 Tabungan Kelas {kelas}")
    render_notif_ultah(kelas)
    df_s = ambil_siswa(kelas)
    if df_s.empty:
        st.warning("Belum ada siswa.")
        return
    map_s = {r["nama_siswa"]: r for _, r in df_s.iterrows()}
    siswa = map_s[st.selectbox("Siswa", list(map_s.keys()))]
    with st.form("tabungan"):
        jenis = st.selectbox("Jenis", ["Setor", "Tarik"])
        nominal = st.number_input("Nominal", min_value=0, step=1000)
        tgl = st.date_input("Tanggal")
        ket = st.text_input("Keterangan")
        if st.form_submit_button("💾 Simpan", type="primary"):
            nilai = float(nominal) if jenis == "Setor" else -float(nominal)
            simpan_dengan_offline("tabungan_siswa", {
                "siswa_id": int(siswa["id"]), "jenis": jenis, "nominal": nilai,
                "tanggal": tgl.isoformat(), "kelas": kelas, "keterangan": ket,
            }, "Tabungan tersimpan!")
    st.subheader("Rekap — Nama | Nominal")
    r = rekap_tabungan_kelas(kelas)
    r["Nominal"] = r["Nominal"].apply(format_rupiah)
    st.dataframe(r, use_container_width=True, hide_index=True)


def halaman_data_siswa():
    pastikan_akses_halaman("Data Siswa")
    kelas = _kelas_aktif()
    st.header(f"👧 Data Siswa Kelas {kelas}")
    render_notif_ultah(kelas)
    st.download_button("📥 Template Excel", buat_template_excel_siswa(),
                       "Template_Siswa.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    up = st.file_uploader("Unggah Excel", type=["xlsx", "xls"])
    if up and st.button("⬆️ Impor", type="primary"):
        ok, msg = proses_upload_excel_siswa(up.getvalue(), kelas)
        st.success(msg) if ok else st.error(msg)
    df = ambil_siswa(kelas)
    if df.empty:
        st.info("Belum ada data.")
    else:
        kol = [c for c in ["no", "nama_siswa", "nis", "nama_ibu", "alamat", "no_whatsapp",
                           "tanggal_lahir", "biaya_mobil", "biaya_lks"] if c in df.columns]
        st.dataframe(df[kol], use_container_width=True, hide_index=True)
        st.caption(f"Total: **{len(df)}** siswa")


def halaman_input_transaksi_bendahara():
    pastikan_akses_halaman("Input Transaksi Bendahara")
    st.header("💰 Input Transaksi Bendahara")
    daftar_pos = pos_input_tersedia()
    if not daftar_pos:
        st.error("Tidak punya hak input.")
        st.stop()
    if st.session_state.peran_aktif == ROLE_TU_BENDAHARA:
        st.info("Pos dikunci: **DANANTARA**")
    with st.form("input_bendahara", clear_on_submit=True):
        pos = st.selectbox("Pos Dana", daftar_pos, disabled=len(daftar_pos) == 1)
        jenis = st.selectbox("Jenis", ["IN", "OUT"])
        tanggal = st.date_input("Tanggal", value=date.today())
        nominal = st.number_input("Nominal", min_value=0, step=1000)
        ket = st.text_input("Keterangan")
        foto = st.file_uploader("Foto Nota", type=TIPE_FILE_NOTA)
        if st.form_submit_button("💾 Simpan", type="primary") and nominal > 0:
            if not boleh_input_transaksi_bendahara(pos):
                st.error("⛔ Tidak punya izin pos ini.")
            else:
                url_foto = unggah_foto_nota(foto.getvalue(), foto.name, foto.type) if foto and cek_koneksi_internet() else None
                payload = {
                    "nama_bendahara": st.session_state.get("nama_guru") or "Bendahara",
                    "pos_dana": pos, "jenis_transaksi": jenis_transaksi_dari_kode(jenis),
                    "nominal": float(nominal), "tanggal": tanggal.isoformat(),
                    "keterangan": ket, "foto_nota": url_foto,
                }
                if cek_koneksi_internet() and simpan_transaksi_bendahara(**payload):
                    st.session_state.flash_pesan = "Transaksi tersimpan."
                    muat_ulang_aplikasi()
                else:
                    simpan_dengan_offline("transaksi_bendahara", payload, "Transaksi (draft)")


def halaman_rekap_transaksi():
    pastikan_akses_halaman("Rekap Transaksi")
    st.header("📊 Rekap Transaksi")
    if not boleh_lihat_rekap_semua_pos():
        st.error("⛔ Tidak punya akses.")
        st.stop()
    df = ambil_transaksi_bendahara()
    if df.empty:
        st.info("Belum ada transaksi.")
        return
    bln = st.selectbox("Bulan", list(NAMA_BULAN.keys()), format_func=lambda x: NAMA_BULAN[x],
                       index=date.today().month - 1)
    thn = st.selectbox("Tahun", list(range(2024, date.today().year + 2)),
                       index=list(range(2024, date.today().year + 2)).index(date.today().year))
    df_b = filter_bulan_tahun(df, bln, thn)
    ringkas = ringkasan_bendahara(df_b)
    lap = laporan_per_pos_dana(df_b)
    c1, c2, c3 = st.columns(3)
    c1.metric("Pemasukan", format_rupiah(ringkas["pemasukan"]))
    c2.metric("Pengeluaran", format_rupiah(ringkas["pengeluaran"]))
    c3.metric("Saldo", format_rupiah(ringkas["saldo"]))
    peran = st.session_state.peran_aktif
    if peran == ROLE_GURU_BENDAHARA and not df_b.empty:
        st.download_button("⬇️ Excel Semua Pos", buat_excel_laporan_pos(bln, thn, ringkas, lap),
                           f"Laporan_{NAMA_BULAN[bln]}_{thn}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    elif peran == ROLE_TU_BENDAHARA:
        df_d = df_b[df_b["pos_dana"] == "DANANTARA"]
        if not df_d.empty:
            st.download_button("⬇️ Excel DANANTARA", dataframe_to_excel_bytes(df_d),
                               f"DANANTARA_{bln}_{thn}.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    for pos, data in lap.items():
        with st.expander(f"📁 {pos}"):
            d = data["detail"].copy()
            d["tanggal"] = d["tanggal"].dt.strftime("%d/%m/%Y")
            d["nominal"] = d["nominal"].apply(format_rupiah)
            st.dataframe(d[["tanggal", "nama_bendahara", "jenis_transaksi", "nominal", "keterangan"]],
                         use_container_width=True, hide_index=True)


def halaman_dashboard_eksekutif():
    pastikan_akses_halaman("Dashboard Eksekutif")
    st.header("📈 Dashboard Eksekutif")
    st.caption("View-Only — Kepsek & Komite")
    metrik = metrik_pelunasan_sekolah(date.today().month, date.today().year)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Saldo Konsolidasi", format_rupiah(hitung_saldo_konsolidasi_sekolah()))
    c2.metric("Total Pemasukan", format_rupiah(hitung_total_pemasukan()))
    c3.metric("Total Pengeluaran", format_rupiah(hitung_total_pengeluaran()))
    c4.metric("Capaian Iuran", f"{metrik['persen_sekolah']}%")
    st.subheader("Saldo per Pos Dana")
    t = saldo_per_pos_dana()
    if not t.empty:
        for kol in ("Pemasukan", "Pengeluaran", "Saldo"):
            t[kol] = t[kol].apply(format_rupiah)
        st.dataframe(t, use_container_width=True, hide_index=True)
    df_k = metrik["per_kelas"]
    if not df_k.empty:
        fig = px.bar(df_k, x="Kelas", y="Persentase", text="Persentase", color="Persentase",
                     color_continuous_scale="Greens")
        fig.update_traces(texttemplate="%{text:.1f}%", textposition="outside")
        fig.update_layout(yaxis_range=[0, 110])
        st.plotly_chart(fig, use_container_width=True)


ROUTER = {
    "Absen Kelas": halaman_absen_kelas,
    "Iuran Kelas": halaman_iuran_kelas,
    "Tabungan Kelas": halaman_tabungan_kelas,
    "Data Siswa": halaman_data_siswa,
    "Input Transaksi Bendahara": halaman_input_transaksi_bendahara,
    "Rekap Transaksi": halaman_rekap_transaksi,
    "Dashboard Eksekutif": halaman_dashboard_eksekutif,
}


# =============================================================================
# MAIN
# =============================================================================

def main():
    st.set_page_config(page_title="Keuangan MI Tempelsari", page_icon="🏫",
                       layout="wide", initial_sidebar_state="expanded")
    init_auth_state()
    init_offline_state()
    gabung_antrean_dari_url()
    coba_auto_login_dari_remember()
    pastikan_bucket_storage()
    inject_css_global()

    with st.sidebar:
        tampilkan_logo_sidebar()
        st.title("🏫 MI Tempelsari")
        st.caption("Keuangan Terpadu")
        st.markdown("---")

        if not st.session_state.authenticated:
            st.markdown("##### 🔐 Login")
            username_input = st.text_input("Username", key="login_id", placeholder="username dari daftar_guru")
            password = st.text_input("Password", type="password", key="login_pw")
            ingat = st.checkbox("Ingat Saya (auto-login)", key="login_ingat")
            if st.button("🔓 Masuk", use_container_width=True, type="primary"):
                ok, pesan = proses_login_daftar_guru(username_input, password, ingat)
                if ok:
                    st.session_state.flash_pesan = pesan
                    muat_ulang_aplikasi()
                else:
                    st.error(pesan)
            st.info("Login: kolom `username` + `password` dari tabel `daftar_guru`.")
            st.stop()

        peran = st.session_state.peran_aktif
        nama = st.session_state.username or "Pengguna"
        st.success(f"**{nama}**")
        st.caption(f"Role: **{ROLE_LABEL.get(peran, peran)}** (`{peran}`)")
        if st.session_state.remember_me:
            st.caption("✅ Auto-login aktif")

        if punya_akses_kelas():
            if st.session_state.kelas_guru:
                st.caption(f"📚 Kelas diampu: **{st.session_state.kelas_guru}**")
            else:
                st.warning("Kolom `kelas_diampu` kosong di database.")
        if peran == ROLE_GURU_BENDAHARA:
            st.caption("✅ Akses: menu kelas + input transaksi bendahara")
            if st.session_state.pos_bendahara:
                st.caption("💰 Pos: " + ", ".join(st.session_state.pos_bendahara))
        if peran == ROLE_KEPSEK_KOMITE:
            st.caption("📈 Akses: dashboard monitoring (view-only)")
        if peran == ROLE_TU_BENDAHARA:
            st.caption("💰 Input: **DANANTARA**")

        render_form_ubah_password()
        st.markdown("---")
        render_monitor_koneksi_offline()
        n = jumlah_antrean_offline()
        if n > 0:
            st.caption(f"📦 **{n}** draft kuning")
            if st.button("🔄 Sinkronkan Data Offline", use_container_width=True, type="primary"):
                if cek_koneksi_internet():
                    ok_n, fail_n = sinkronkan_antrean_offline()
                    if fail_n == 0 and ok_n > 0:
                        bersihkan_localstorage_browser()
                        st.session_state.flash_pesan = f"✅ {ok_n} draft disinkronkan!"
                        st.session_state.offline_draft_baru = False
                        muat_ulang_aplikasi()
                    elif ok_n:
                        st.warning(f"{ok_n} ok, {fail_n} gagal.")
                    else:
                        st.warning("Sinkronisasi gagal.")
                else:
                    st.warning("Masih offline.")
        if st.button("🚪 Logout", use_container_width=True):
            proses_logout()
        st.markdown("---")
        menu = menu_berdasarkan_peran(peran)
        if st.session_state.halaman_aktif not in menu:
            st.session_state.halaman_aktif = menu[0]
        st.session_state.halaman_aktif = st.radio("Menu", menu,
            index=menu.index(st.session_state.halaman_aktif), label_visibility="collapsed")
        if st.button("🔄 Muat Ulang", use_container_width=True):
            muat_ulang_aplikasi()

    _l, _c, _r = st.columns([2, 1, 2])
    with _c:
        tampilkan_logo_utama()
    st.title("🏫 Aplikasi Keuangan Terpadu MI Al Ma'arif Tempelsari")
    st.markdown('<p class="sub-judul">RBAC via daftar_guru · Streamlit + Supabase</p>', unsafe_allow_html=True)
    if st.session_state.flash_pesan:
        st.success(st.session_state.flash_pesan)
        st.session_state.flash_pesan = None
    render_banner_draft_kuning()
    render_grid_menu_utama()
    st.markdown("---")
    handler = ROUTER.get(st.session_state.halaman_aktif)
    if handler:
        handler()
    st.markdown("---")
    st.caption("© 2026 MI Al Ma'arif Tempelsari")


if __name__ == "__main__":
    main()
